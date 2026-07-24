"""
Export tools - export test cases to JSON, Excel, Markdown, and TestRail CSV.
All tools read test cases from a file, NOT from a parameter string.
AI should first save test cases with write_file, then call these with just the path.
"""

import json
import os
from agent.tool_registry import tool


def _load_cases(path: str):
    """Load test cases from a JSON/JSONL file. Returns (cases, error)."""
    if not os.path.exists(path):
        return None, f"File not found: {path}. First save test cases with write_file."
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if not content:
            return [], None

        # JSON array: [{...}, {...}]
        if content.startswith("["):
            data = json.loads(content)
            if isinstance(data, list):
                return data, None
            if isinstance(data, dict) and "test_cases" in data:
                return data["test_cases"], None
            return data, None

        # JSON wrapper or JSONL
        if content.startswith("{"):
            try:
                data = json.loads(content)
            except json.JSONDecodeError:
                # Not a single JSON → try JSONL
                data = None

            if isinstance(data, dict) and "test_cases" in data:
                return data["test_cases"], None
            if isinstance(data, dict):
                return [data], None
            if isinstance(data, list):
                return data, None

            # JSONL: one JSON object per line
            cases = []
            for line in content.split("\n"):
                line = line.strip()
                if line:
                    try:
                        cases.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
            return cases, None

        return [], None

    except Exception as e:
        return None, f"Error reading {path}: {e}"


# @tool(  # 不注册到工具箱，仅由前端直接调用
#     name="export_json",
#     description="Export test cases to a formatted JSON file.",
#     parameters={
#         "type": "object",
#         "properties": {
#             "path": {"type": "string", "description": "Source JSON file path"},
#             "output": {"type": "string", "description": "Output path"},
#         },
#         "required": ["path"],
#     },
#     category="output",
# )
def export_json(path: str, output: str = "") -> str:
    cases, err = _load_cases(path)
    if err:
        return err

    out_path = output or path
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    export_data = {
        "export_info": {
            "total": len(cases),
            "exported_at": __import__("datetime").datetime.now().isoformat(),
        },
        "test_cases": cases,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)
    return f"Exported {len(cases)} test cases to {out_path}"


@tool(
    name="export_excel",
    description="Export test cases to an Excel (.xlsx) file. Reads cases from a JSON file (saved by write_file first).",
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "Source JSON file path (e.g. projects/<name>/cases/test_cases.jsonl). Use write_file first to create this file."
            },
            "output": {
                "type": "string",
                "description": "Output Excel path (e.g. cases/test_cases.xlsx)"
            },
        },
        "required": ["path", "output"],
    },
    category="output",
)
def export_excel(path: str, output: str) -> str:
    cases, err = _load_cases(path)
    if err:
        return err

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E2D3D", end_color="1E2D3D", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["用例ID", "所属模块", "测试类型", "优先级", "用例标题", "前置条件", "测试步骤", "预期结果", "数据类型", "标签"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    priority_fills = {
        "P0": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "P1": PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),
        "P2": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        "P3": PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid"),
    }

    for row, tc in enumerate(cases, 2):
        steps_text = ""
        for s in tc.get("steps", []):
            a = s.get("action", "") if isinstance(s, dict) else str(s)
            d = s.get("data", "") if isinstance(s, dict) else ""
            e = s.get("expected", "") if isinstance(s, dict) else ""
            parts = [f"Step {s.get('step', '?')}: {a}" if isinstance(s, dict) else a]
            if d:
                parts.append(f"Data: {d}")
            if e:
                parts.append(f"Expected: {e}")
            steps_text += " | ".join(parts) + "\n"

        values = [
            tc.get("id", ""), tc.get("module", ""), tc.get("type", "API"),
            tc.get("priority", "P2"), tc.get("title", ""), tc.get("precondition", ""),
            steps_text.strip(), tc.get("expectedResult", ""),
            tc.get("testDataType", ""), ", ".join(tc.get("tags", [])),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        p = tc.get("priority", "P2")
        if p in priority_fills:
            ws.cell(row=row, column=4).fill = priority_fills[p]

    widths = [12, 14, 8, 8, 35, 20, 50, 30, 14, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
    wb.save(output)
    return f"Exported {len(cases)} test cases to {output}"


# @tool(  # 不注册到工具箱，仅由前端直接调用
#     name="export_markdown",
#     description="Export test cases as a Markdown report.",
#     parameters={
#         "type": "object",
#         "properties": {
#             "path": {"type": "string", "description": "Source JSON file path"},
#             "output": {"type": "string", "description": "Output Markdown path"},
#             "title": {"type": "string", "description": "Report title"},
#         },
#         "required": ["path", "output"],
#     },
#     category="output",
# )
def export_markdown(path: str, output: str, title: str = "测试用例报告") -> str:
    cases, err = _load_cases(path)
    if err:
        return err

    lines = [
        f"# {title}", "",
        f"**总计**: {len(cases)} 条用例  ",
        f"**生成时间**: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        "", "## 按优先级统计", "",
        "| 优先级 | 数量 |", "|--------|------|",
    ]
    pcount = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for tc in cases:
        pcount[tc.get("priority", "P2")] = pcount.get(tc.get("priority", "P2"), 0) + 1
    for p in ["P0", "P1", "P2", "P3"]:
        lines.append(f"| {p} | {pcount.get(p, 0)} |")

    lines += ["", "## 按模块统计", "", "| 模块 | 数量 |", "|------|------|"]
    mcount = {}
    for tc in cases:
        m = tc.get("module", "未知")
        mcount[m] = mcount.get(m, 0) + 1
    for m, c in sorted(mcount.items()):
        lines.append(f"| {m} | {c} |")

    lines += ["", "## 用例详情", ""]
    for tc in cases:
        lines.append(f"### {tc.get('id', '?')}: {tc.get('title', '')}")
        lines.append(f"- **所属模块**: {tc.get('module', '')}")
        lines.append(f"- **测试类型**: {tc.get('type', 'API')}  ")
        lines.append(f"- **优先级**: {tc.get('priority', 'P2')}  ")
        lines.append(f"- **数据类型**: {tc.get('testDataType', '')}  ")
        precond = tc.get("precondition", "")
        if precond:
            lines.append(f"- **前置条件**: {precond}  ")
        lines.append(f"- **预期结果**: {tc.get('expectedResult', '')}  ")
        steps = tc.get("steps", [])
        if steps:
            lines.append("")
            lines.append("| 步骤 | 操作 | 数据 | 预期 |")
            lines.append("|------|------|------|------|")
            for s in steps:
                if isinstance(s, dict):
                    lines.append(f"| {s.get('step', '?')} | {s.get('action', '')} | {s.get('data', '') or '-'} | {s.get('expected', '') or '-'} |")
        lines.append("")

    report = "\n".join(lines)
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
    with open(output, "w", encoding="utf-8") as f:
        f.write(report)
    return f"Exported {len(cases)} test cases to {output}"


# @tool(  # 不注册到工具箱，仅由前端直接调用
#     name="export_pdf",
#     description="Export test cases to a PDF report.",
#     parameters={
#         "type": "object",
#         "properties": {
#             "path": {"type": "string", "description": "Source JSON file path"},
#             "output": {"type": "string", "description": "Output PDF path"},
#             "title": {"type": "string", "description": "Report title"},
#         },
#         "required": ["path", "output"],
#     },
#     category="output",
# )
def export_pdf(path: str, output: str, title: str = "测试用例报告") -> str:
    cases, err = _load_cases(path)
    if err:
        return err

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.lib.colors import HexColor, white
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return "错误: reportlab 未安装。运行: pip install reportlab"

    # --- Font setup: use CID font (macOS PingFang.ttc uses postscript outlines, unsupported by reportlab) ---
    font_name = "STSong-Light"
    font_bold = "STSong-Light"

    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    except Exception:
        font_name = "Helvetica"
        font_bold = "Helvetica-Bold"

    doc = SimpleDocTemplate(
        output, pagesize=A4,
        topMargin=2*cm, bottomMargin=1.5*cm,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
    )

    s = getSampleStyleSheet()

    def _style(name, parent_name, **kw):
        base = s[parent_name]
        kw.setdefault("fontName", font_name)
        return ParagraphStyle(name, parent=base, **kw)

    st = {
        "title": _style("PdfTitle", "Title", fontName=font_bold, fontSize=18, leading=24,
                        textColor=HexColor("#1e293b"), spaceAfter=4*mm),
        "subtitle": _style("PdfSub", "Normal", fontSize=10, leading=14,
                           textColor=HexColor("#64748b"), spaceAfter=8*mm),
        "section": _style("PdfSection", "Heading2", fontName=font_bold, fontSize=13, leading=18,
                          textColor=HexColor("#334155"), spaceBefore=6*mm, spaceAfter=3*mm),
        "body": _style("PdfBody", "Normal", fontSize=9, leading=13,
                       textColor=HexColor("#334155"), spaceAfter=1*mm),
        "case_title": _style("PdfCase", "Heading3", fontName=font_bold, fontSize=11, leading=15,
                             textColor=HexColor("#1e293b"), spaceBefore=3*mm, spaceAfter=2*mm),
        "case_sep": _style("CaseSep", "Normal", fontSize=6, leading=6,
                           textColor=HexColor("#cbd5e1")),
    }

    elements = []

    # Title
    elements.append(Paragraph(title, st["title"]))
    now_str = __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")
    elements.append(Paragraph(f"共 {len(cases)} 条用例  |  生成时间: {now_str}", st["subtitle"]))

    # Summary
    pcount = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    mcount = {}
    for tc in cases:
        pcount[tc.get("priority", "P2")] = pcount.get(tc.get("priority", "P2"), 0) + 1
        mcount[tc.get("module", "未知")] = mcount.get(tc.get("module", "未知"), 0) + 1

    summary_data = [["优先级", "P0", "P1", "P2", "P3", "总计"],
                    ["数量", str(pcount["P0"]), str(pcount["P1"]),
                     str(pcount["P2"]), str(pcount["P3"]), str(len(cases))]]
    pri_colors = {"P0": HexColor("#fecaca"), "P1": HexColor("#fed7aa"),
                  "P2": HexColor("#fef08a"), "P3": HexColor("#bbf7d0")}
    summary_table = Table(summary_data, colWidths=[40*mm, 25*mm, 25*mm, 25*mm, 25*mm, 25*mm])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (1, 1), (1, 1), pri_colors["P0"]),
        ("BACKGROUND", (2, 1), (2, 1), pri_colors["P1"]),
        ("BACKGROUND", (3, 1), (3, 1), pri_colors["P2"]),
        ("BACKGROUND", (4, 1), (4, 1), pri_colors["P3"]),
    ]))
    elements.append(Paragraph("按优先级统计", st["section"]))
    elements.append(summary_table)

    # Module distribution
    mod_data = [["模块", "用例数"]]
    for m, c in sorted(mcount.items(), key=lambda x: -x[1]):
        mod_data.append([m, str(c)])
    mod_table = Table(mod_data, colWidths=[90*mm, 50*mm])
    mod_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, HexColor("#f8fafc")]),
    ]))
    elements.append(Paragraph("按模块统计", st["section"]))
    elements.append(mod_table)

    # Cases detail
    elements.append(Paragraph("用例详情", st["section"]))
    elements.append(HRFlowable(width="100%", thickness=0.5, lineCap="round",
                                color=HexColor("#cbd5e1"), spaceBefore=2*mm, spaceAfter=2*mm))

    for tc in cases:
        tid = tc.get("id", "?")
        ttitle = tc.get("title", "")
        elements.append(Paragraph(f"{tid}: {ttitle}", st["case_title"]))

        mod = tc.get("module", "")
        ttype = tc.get("type", "API")
        pri = tc.get("priority", "P2")
        elements.append(Paragraph(f"所属模块: {mod}  |  类型: {ttype}  |  优先级: {pri}", st["body"]))

        if tc.get("precondition"):
            elements.append(Paragraph(f"前置条件: {tc['precondition']}", st["body"]))
        if tc.get("expectedResult"):
            elements.append(Paragraph(f"预期结果: {tc['expectedResult']}", st["body"]))

    # Cell paragraph styles
    _cell_style = ParagraphStyle("PdfCell", parent=st["body"], fontSize=7, leading=10)
    _hdr_style = ParagraphStyle("PdfCellHdr", parent=st["body"], fontSize=7, leading=10,
                                fontName=font_bold, textColor=HexColor("#334155"))

    for tc in cases:
        tid = tc.get("id", "?")
        ttitle = tc.get("title", "")
        elements.append(Paragraph(f"{tid}: {ttitle}", st["case_title"]))

        mod = tc.get("module", "")
        ttype = tc.get("type", "API")
        pri = tc.get("priority", "P2")
        elements.append(Paragraph(f"所属模块: {mod}  |  类型: {ttype}  |  优先级: {pri}", st["body"]))

        if tc.get("precondition"):
            elements.append(Paragraph(f"前置条件: {tc['precondition']}", st["body"]))
        if tc.get("expectedResult"):
            elements.append(Paragraph(f"预期结果: {tc['expectedResult']}", st["body"]))

        steps = tc.get("steps", [])
        if steps:
            def _esc(v):
                """Convert value to HTML-safe string for Paragraph."""
                if isinstance(v, dict):
                    raw = json.dumps(v, ensure_ascii=False)
                elif v is None:
                    raw = "-"
                else:
                    raw = str(v)
                return raw.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

            step_data = [[Paragraph("步骤", _hdr_style),
                          Paragraph("操作", _hdr_style),
                          Paragraph("数据", _hdr_style),
                          Paragraph("预期", _hdr_style)]]
            for s in steps:
                if isinstance(s, dict):
                    step_data.append([
                        Paragraph(_esc(s.get("step")), _cell_style),
                        Paragraph(_esc(s.get("action")), _cell_style),
                        Paragraph(_esc(s.get("data")), _cell_style),
                        Paragraph(_esc(s.get("expected")), _cell_style),
                    ])
            step_table = Table(step_data, colWidths=[12*mm, 45*mm, 35*mm, 45*mm])
            step_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#334155")),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(Spacer(1, 2*mm))
            elements.append(step_table)

        elements.append(Spacer(1, 3*mm))
        elements.append(HRFlowable(width="100%", thickness=0.3, lineCap="round",
                                    color=HexColor("#e2e8f0"), spaceBefore=1*mm, spaceAfter=1*mm))

    doc.build(elements)
    return f"已导出 {len(cases)} 条用例到 {output}"


# TODO: TestRail CSV 导出已保留但暂不注册到工具列表，等配置对接后启用
# @tool(
#     name="export_testrail_csv",
#     description="Export test cases to TestRail-compatible CSV. Reads cases from a JSON file (saved by write_file first).",
#     parameters={
#         "type": "object",
#         "properties": {
#             "path": {
#                 "type": "string",
#                 "description": "Source JSON file path (e.g. projects/<name>/cases/test_cases.jsonl). Use write_file first to create this file."
#             },
#             "output": {
#                 "type": "string",
#                 "description": "Output CSV path (e.g. cases/testrail.csv)"
#             },
#         },
#         "required": ["path", "output"],
#     },
#     category="output",
# )
def export_testrail_csv(path: str, output: str) -> str:
    cases, err = _load_cases(path)
    if err:
        return err

    import csv
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)

    fieldnames = ["ID", "Title", "Section", "Type", "Priority", "Preconditions", "Steps", "Expected Result", "Automation Type", "References"]
    with open(output, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for tc in cases:
            steps_text = ""
            for s in tc.get("steps", []):
                if isinstance(s, dict):
                    steps_text += f"Step {s.get('step', '?')}: {s.get('action', '')}\n"
                    if s.get("data"):
                        steps_text += f"  Data: {s.get('data')}\n"
                    if s.get("expected"):
                        steps_text += f"  Expected: {s.get('expected')}\n"
                else:
                    steps_text += str(s) + "\n"

            writer.writerow({
                "ID": tc.get("id", ""),
                "Title": tc.get("title", ""),
                "Section": tc.get("module", ""),
                "Type": "Functional" if tc.get("type") == "API" else tc.get("type", "Functional"),
                "Priority": {"P0": "Critical", "P1": "High", "P2": "Medium", "P3": "Low"}.get(tc.get("priority", "P2"), "Medium"),
                "Preconditions": tc.get("precondition", ""),
                "Steps": steps_text.strip(),
                "Expected Result": tc.get("expectedResult", ""),
                "Automation Type": "Manual",
                "References": ", ".join(tc.get("tags", [])),
            })

    return f"Exported {len(cases)} test cases to {output} (TestRail CSV)"


def export_csv(path: str, output: str, title: str = "测试用例报告") -> str:
    """导出测试用例为通用 CSV 文件（中文表头）。"""
    cases, err = _load_cases(path)
    if err:
        return err

    import csv
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)

    fieldnames = ["编号", "标题", "模块", "优先级", "测试类型", "前置条件", "测试步骤", "预期结果", "标签"]
    with open(output, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for tc in cases:
            steps_text = ""
            for s in tc.get("steps", []):
                if isinstance(s, dict):
                    parts = [str(s.get("step", "")), str(s.get("action", ""))]
                    if s.get("data"):
                        parts.append(str(s["data"]))
                    parts.append(str(s.get("expected", "")))
                    steps_text += " | ".join(p for p in parts if p) + "\n"
                else:
                    steps_text += str(s) + "\n"

            writer.writerow({
                "编号": tc.get("id", ""),
                "标题": tc.get("title", ""),
                "模块": tc.get("module", ""),
                "优先级": tc.get("priority", ""),
                "测试类型": tc.get("type", ""),
                "前置条件": tc.get("precondition", ""),
                "测试步骤": steps_text.strip(),
                "预期结果": tc.get("expectedResult", ""),
                "标签": ", ".join(tc.get("tags", [])),
            })

    return f"已导出 {len(cases)} 条用例到 {output}"
